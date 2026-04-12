#!/usr/bin/env python3
"""
load_json_output_into_xlsx.py

Populate the SQLite-design XLSX (acting as a staging table set) from a folder of
Schema.org JSON-LD files (json_output).

Improvements over prior version:
- INTERNAL link resolution is robust:
    * supports relative links (/path, path)
    * strips fragments
    * normalises scheme/host/path
    * resolves by URL, by WebPage @id URI, and by path fallback
- INTERNAL links no longer pollute ENT_external_reference:
    * if internal but unresolved -> leave destination columns blank and report
- Skips writing junk rows (e.g., blank webpage_id in junctions)
- Enforces "exactly one destination column" for resolved links
- Optional link-resolution report CSV (useful in CI)

Designed to run in a locked-down GitHub Actions environment with only openpyxl installed.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlparse, urlunparse, urljoin

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


INTERNAL_DOMAIN = "ipfirstresponse.ipaustralia.gov.au"
INTERNAL_BASE = f"https://{INTERNAL_DOMAIN}"


# -----------------------------
# Helpers
# -----------------------------

def norm_url(u: str) -> str:
    """
    Normalise URLs for matching:
    - strip whitespace
    - drop fragments (#...)
    - lower-case scheme + netloc
    - trim trailing slash for non-root paths
    """
    if not u:
        return ""
    u = str(u).strip()
    if not u:
        return ""

    parsed = urlparse(u)
    parsed = parsed._replace(fragment="")

    scheme = (parsed.scheme or "").lower()
    netloc = (parsed.netloc or "").lower()

    # Rebuild first (keeping query as-is; we may drop later for internal matching)
    rebuilt = urlunparse(parsed._replace(scheme=scheme, netloc=netloc))

    # Trim trailing slash unless root
    p2 = urlparse(rebuilt)
    if rebuilt.endswith("/") and (p2.path and p2.path != "/"):
        rebuilt = rebuilt.rstrip("/")

    return rebuilt


def drop_common_tracking_params(u: str) -> str:
    """
    Optional: if internal URLs come with tracking params, treat them as equivalent.
    We keep it simple: if internal, drop *all* query params.
    """
    if not u:
        return ""
    try:
        p = urlparse(u)
        if p.netloc.lower() == INTERNAL_DOMAIN:
            return urlunparse(p._replace(query="", fragment=""))
    except Exception:
        pass
    return u


def is_internal_ipfr_url(u: str) -> bool:
    if not u:
        return False
    try:
        return urlparse(norm_url(u)).netloc.lower() == INTERNAL_DOMAIN
    except Exception:
        return False


def to_internal_absolute(maybe_url: str) -> str:
    """
    Convert relative or scheme-less internal links into absolute IPFR URLs when reasonable.
    Examples:
      /options/foo -> https://ipfirstresponse.ipaustralia.gov.au/options/foo
      options/foo  -> https://ipfirstresponse.ipaustralia.gov.au/options/foo
      //ipfirstresponse.ipaustralia.gov.au/x -> https://ipfirstresponse.ipaustralia.gov.au/x
    """
    if not maybe_url:
        return ""
    raw = str(maybe_url).strip()
    if not raw:
        return ""

    # Skip non-http link types
    lowered = raw.lower()
    if lowered.startswith(("mailto:", "tel:", "javascript:", "#")):
        return ""

    # Scheme-relative URL
    if raw.startswith("//"):
        raw = "https:" + raw

    p = urlparse(raw)

    # Already absolute
    if p.scheme in ("http", "https") and p.netloc:
        return raw

    # Root-relative
    if raw.startswith("/"):
        return urljoin(INTERNAL_BASE, raw)

    # Bare path (no scheme, no netloc) – treat as internal path
    if not p.scheme and not p.netloc and p.path:
        return urljoin(INTERNAL_BASE + "/", raw)

    return raw


def sha1_id(text: str, prefix: str, length: int = 8) -> str:
    h = hashlib.sha1(text.encode("utf-8")).hexdigest()[:length]
    return f"{prefix}{h}"


def clear_sheet(ws: Worksheet) -> None:
    """Delete all rows except header."""
    if ws.max_row and ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def headers(ws: Worksheet) -> List[str]:
    return [c.value for c in ws[1] if c.value is not None]


def append_row(ws: Worksheet, header_list: List[str], row: Dict[str, Any]) -> None:
    ws.append([row.get(h, None) for h in header_list])


def first(it: List[Any]) -> Optional[Any]:
    return it[0] if it else None


def safe_list(x: Any) -> List[Any]:
    if x is None:
        return []
    if isinstance(x, list):
        return x
    return [x]


# -----------------------------
# Registries
# -----------------------------

@dataclass
class Registries:
    stakeholder_uri_to_id: Dict[str, str]
    website_uri_to_id: Dict[str, str]
    external_url_to_id: Dict[str, str]
    term_key_to_id: Dict[Tuple[str, str], str]  # (term_to_define, wikidata_url)
    disclaimer_text_to_id: Dict[str, str]       # normalised text -> disclaimer_id

    stakeholder_next: int = 1
    website_next: int = 1
    term_next: int = 1
    extref_next: int = 1


def stakeholder_id_for(reg: Registries, uri: str) -> str:
    uri = uri or ""
    if uri in reg.stakeholder_uri_to_id:
        return reg.stakeholder_uri_to_id[uri]
    sid = f"SH_{reg.stakeholder_next:04d}"
    reg.stakeholder_next += 1
    reg.stakeholder_uri_to_id[uri] = sid
    return sid


def website_id_for(reg: Registries, uri: str) -> str:
    uri = uri or ""
    if uri in reg.website_uri_to_id:
        return reg.website_uri_to_id[uri]
    wid = f"SITE_{reg.website_next:04d}"
    reg.website_next += 1
    reg.website_uri_to_id[uri] = wid
    return wid


def external_ref_id_for(reg: Registries, url: str) -> str:
    url = norm_url(url)
    if url in reg.external_url_to_id:
        return reg.external_url_to_id[url]
    eid = f"extref_{reg.extref_next:04d}"
    reg.extref_next += 1
    reg.external_url_to_id[url] = eid
    return eid


def term_id_for(reg: Registries, term: str, wikidata_url: str) -> str:
    key = (term or "", wikidata_url or "")
    if key in reg.term_key_to_id:
        return reg.term_key_to_id[key]
    tid = f"T{reg.term_next:04d}"
    reg.term_next += 1
    reg.term_key_to_id[key] = tid
    return tid


def disclaimer_id_for(reg: Registries, text: str) -> str:
    norm = " ".join(str(text or "").strip().split())
    if not norm:
        return ""
    if norm in reg.disclaimer_text_to_id:
        return reg.disclaimer_text_to_id[norm]
    did = sha1_id(norm, "DCL_")
    reg.disclaimer_text_to_id[norm] = did
    return did


# -----------------------------
# JSON utilities
# -----------------------------

def load_json(path: str) -> Optional[dict]:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def graph_nodes(doc: dict) -> List[dict]:
    if not isinstance(doc, dict):
        return []
    g = doc.get("@graph")
    if isinstance(g, list):
        return [n for n in g if isinstance(n, dict)]
    return []


def find_nodes_by_type(nodes: List[dict], t: str) -> List[dict]:
    return [n for n in nodes if n.get("@type") == t]


def find_node_by_id(nodes: List[dict], node_id: str) -> Optional[dict]:
    for n in nodes:
        if n.get("@id") == node_id:
            return n
    return None


def extract_webpages_with_identifier(nodes: List[dict]) -> List[dict]:
    return [n for n in nodes if n.get("@type") == "WebPage" and n.get("identifier")]


def extract_website(nodes: List[dict]) -> Optional[dict]:
    return first(find_nodes_by_type(nodes, "WebSite"))


# -----------------------------
# Internal resolvers
# -----------------------------

@dataclass
class InternalResolvers:
    url_to_webpage_id: Dict[str, str]
    uri_to_webpage_id: Dict[str, str]
    path_to_webpage_id: Dict[str, str]


def build_internal_resolvers(json_files: List[str]) -> InternalResolvers:
    """
    Scan all JSONs once, mapping:
    - webpage_url -> webpage_id (identifier)
    - webpage_uri (@id) -> webpage_id
    - path (/foo/bar) -> webpage_id (internal only), as fallback
    """
    url_map: Dict[str, str] = {}
    uri_map: Dict[str, str] = {}
    path_map: Dict[str, str] = {}

    for p in json_files:
        doc = load_json(p)
        if not doc:
            continue

        for wp in extract_webpages_with_identifier(graph_nodes(doc)):
            wp_id = str(wp.get("identifier")).strip()
            if not wp_id:
                continue

            wurl_raw = wp.get("url") or ""
            wurl = norm_url(to_internal_absolute(str(wurl_raw)))
            wurl = norm_url(drop_common_tracking_params(wurl))

            wuri = norm_url(str(wp.get("@id") or ""))

            if wurl:
                url_map[wurl] = wp_id
                url_map[wurl + "/"] = wp_id  # tolerate slash variants
                # path fallback
                try:
                    pu = urlparse(wurl)
                    if pu.netloc.lower() == INTERNAL_DOMAIN and pu.path:
                        path_map[pu.path.rstrip("/") or "/"] = wp_id
                except Exception:
                    pass

            if wuri:
                uri_map[wuri] = wp_id

    return InternalResolvers(url_to_webpage_id=url_map, uri_to_webpage_id=uri_map, path_to_webpage_id=path_map)


def resolve_internal_destination_id(res: InternalResolvers, raw_out: str) -> Optional[str]:
    """
    Resolve an internal outbound link to a webpage_id using:
      1) url_to_webpage_id (after absolute+normalise)
      2) uri_to_webpage_id (if the link is actually a JSON @id style)
      3) path_to_webpage_id fallback
    """
    if not raw_out:
        return None

    abs_u = to_internal_absolute(raw_out)
    if not abs_u:
        return None

    u = norm_url(abs_u)
    u = norm_url(drop_common_tracking_params(u))

    # 1) direct URL match
    if u in res.url_to_webpage_id:
        return res.url_to_webpage_id[u]
    if (u + "/") in res.url_to_webpage_id:
        return res.url_to_webpage_id[u + "/"]

    # 2) treat as @id / uri match
    if u in res.uri_to_webpage_id:
        return res.uri_to_webpage_id[u]

    # 3) path fallback (internal only)
    try:
        pu = urlparse(u)
        if pu.netloc.lower() == INTERNAL_DOMAIN:
            path = (pu.path or "/").rstrip("/") or "/"
            if path in res.path_to_webpage_id:
                return res.path_to_webpage_id[path]
    except Exception:
        pass

    return None


# -----------------------------
# Main
# -----------------------------

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--json_dir", required=True, help="Folder containing json_output files")
    ap.add_argument("--xlsx_path", required=True, help="Path to the XLSX to populate (in-place)")
    ap.add_argument(
        "--link_report_path",
        default="",
        help="Optional CSV output path for unresolved internal links and summary (recommended in CI).",
    )
    args = ap.parse_args()

    json_dir = args.json_dir
    xlsx_path = args.xlsx_path
    report_path = args.link_report_path

    if not os.path.exists(json_dir):
        raise SystemExit(f"JSON directory not found: {json_dir}")
    if not os.path.exists(xlsx_path):
        raise SystemExit(f"XLSX file not found: {xlsx_path}")

    json_files = [
        os.path.join(json_dir, f)
        for f in os.listdir(json_dir)
        if f.lower().endswith(".json")
    ]
    json_files.sort()

    resolvers = build_internal_resolvers(json_files)

    wb = load_workbook(xlsx_path)

    required = [
        "ENT_website",
        "ENT_webpage",
        "ENT_stakeholder",
        "ENT_external_reference",
        "JUNCT_webpage_links",
        "ENT_webpage_main_content",
        "ENT_webpage_faq",
        "ENT_disclaimer",
        "JUNCT_disclaimer",
        "ENT_term",
        "JUNCT_webpage_relevant_terms",
        "JUNCT_term_alias",
    ]
    missing = [s for s in required if s not in wb.sheetnames]
    if missing:
        raise SystemExit(f"Missing sheets in workbook: {missing}")

    ws_website = wb["ENT_website"]
    ws_webpage = wb["ENT_webpage"]
    ws_stakeholder = wb["ENT_stakeholder"]
    ws_extref = wb["ENT_external_reference"]
    ws_links = wb["JUNCT_webpage_links"]
    ws_main = wb["ENT_webpage_main_content"]
    ws_faq = wb["ENT_webpage_faq"]
    ws_disclaimer = wb["ENT_disclaimer"]
    ws_jdisclaimer = wb["JUNCT_disclaimer"]
    ws_term = wb["ENT_term"]
    ws_wprel = wb["JUNCT_webpage_relevant_terms"]
    ws_alias = wb["JUNCT_term_alias"]

    # Clear target sheets (keep header)
    for ws in [
        ws_website, ws_webpage, ws_stakeholder, ws_extref, ws_links,
        ws_main, ws_faq, ws_disclaimer, ws_jdisclaimer, ws_term, ws_wprel, ws_alias
    ]:
        clear_sheet(ws)

    H_WEBSITE = headers(ws_website)
    H_WEBPAGE = headers(ws_webpage)
    H_STAKEHOLDER = headers(ws_stakeholder)
    H_EXTREF = headers(ws_extref)
    H_LINKS = headers(ws_links)
    H_MAIN = headers(ws_main)
    H_FAQ = headers(ws_faq)
    H_DISCLAIMER = headers(ws_disclaimer)
    H_JDISCLAIM = headers(ws_jdisclaimer)
    H_TERM = headers(ws_term)
    H_WPREL = headers(ws_wprel)
    H_ALIAS = headers(ws_alias)

    reg = Registries(
        stakeholder_uri_to_id={},
        website_uri_to_id={},
        external_url_to_id={},
        term_key_to_id={},
        disclaimer_text_to_id={},
    )

    # Accumulators
    stakeholders_out: Dict[str, Dict[str, Any]] = {}
    websites_out: Dict[str, Dict[str, Any]] = {}
    webpages_out: Dict[str, Dict[str, Any]] = {}
    extrefs_out: Dict[str, Dict[str, Any]] = {}
    disclaimers_out: Dict[str, Dict[str, Any]] = {}
    terms_out: Dict[str, Dict[str, Any]] = {}

    wp_term_rows: List[Dict[str, Any]] = []
    link_rows: List[Dict[str, Any]] = []
    main_rows: List[Dict[str, Any]] = []
    faq_rows: List[Dict[str, Any]] = []
    junct_disclaimer_rows: List[Dict[str, Any]] = []

    # Link resolution reporting
    unresolved_internal: List[Tuple[str, str]] = []  # (source_webpage_id, url)

    def upsert_stakeholder(node: dict) -> Optional[str]:
        if not node or not isinstance(node, dict):
            return None
        uri = node.get("@id") or ""
        sid = stakeholder_id_for(reg, uri)

        stype = node.get("@type") or ""
        name = node.get("name") or ""
        alt = node.get("alternateName")
        if isinstance(alt, list):
            alt = ", ".join([str(a) for a in alt if a is not None])
        same_as = node.get("sameAs")
        if isinstance(same_as, list):
            same_as = ", ".join([str(s) for s in same_as if s is not None])
        desc = node.get("description") or ""
        url = node.get("url") or ""

        stakeholders_out[sid] = {
            "stakeholder_id": sid,
            "stakeholder_uri": uri,
            "stakeholder_type": stype,
            "stakeholder_name": name,
            "stakeholder_alias": alt,
            "stakeholder_description": desc,
            "stakeholder_url": url,
            "stakeholder_same_as": same_as,
        }
        return sid

    # Process each JSON
    for p in json_files:
        doc = load_json(p)
        if not doc:
            continue
        nodes = graph_nodes(doc)

        # 1) Stakeholders
        for st in find_nodes_by_type(nodes, "GovernmentOrganization") + find_nodes_by_type(nodes, "Organization"):
            upsert_stakeholder(st)

        # 2) Website
        site = extract_website(nodes)
        if site:
            site_uri = site.get("@id") or ""
            site_id = website_id_for(reg, site_uri)

            pub = site.get("publisher") or {}
            pub_idref = pub.get("@id") if isinstance(pub, dict) else None
            owner_sid = stakeholder_id_for(reg, pub_idref or "") if pub_idref else None

            websites_out[site_id] = {
                "website_id": site_id,
                "website_uri": site_uri,
                "website_title": site.get("name") or "",
                "website_url": site.get("url") or "",
                # IMPORTANT: store stakeholder_id (not URI)
                "website_owned_by": owner_sid,
            }

        # 3) WebPages with identifiers
        for wp in extract_webpages_with_identifier(nodes):
            webpage_id = str(wp.get("identifier")).strip()
            if not webpage_id:
                continue

            webpage_uri = wp.get("@id") or ""
            webpage_url = wp.get("url") or ""
            webpage_name = wp.get("name") or ""
            webpage_desc = wp.get("description") or ""

            # publisher -> stakeholder_id
            publisher = wp.get("publisher") or {}
            pub_idref = publisher.get("@id") if isinstance(publisher, dict) else None
            publisher_sid = stakeholder_id_for(reg, pub_idref or "") if pub_idref else None

            # isPartOf -> website_id
            is_part_of = wp.get("isPartOf") or {}
            ipo_idref = is_part_of.get("@id") if isinstance(is_part_of, dict) else None
            website_id = website_id_for(reg, ipo_idref or "") if ipo_idref else None

            # mainEntity -> type and provider
            main_entity = wp.get("mainEntity") or {}
            me_idref = main_entity.get("@id") if isinstance(main_entity, dict) else None

            main_entity_type = ""
            provider_sid: Optional[str] = None

            if me_idref:
                me_node = find_node_by_id(nodes, me_idref)
                if me_node:
                    main_entity_type = me_node.get("@type") or ""
                    if main_entity_type == "Service":
                        prov = me_node.get("provider") or {}
                        prov_idref = prov.get("@id") if isinstance(prov, dict) else None
                        provider_sid = stakeholder_id_for(reg, prov_idref or "") if prov_idref else None
                    elif main_entity_type == "GovernmentService":
                        sop = me_node.get("serviceOperator") or {}
                        sop_idref = sop.get("@id") if isinstance(sop, dict) else None
                        provider_sid = stakeholder_id_for(reg, sop_idref or "") if sop_idref else None
                    else:
                        provider_sid = None

            webpages_out[webpage_id] = {
                "webpage_id": webpage_id,
                "webpage_uri": webpage_uri,
                "webpage_url": webpage_url,
                "webpage_name": webpage_name,
                "webpage_altname": wp.get("alternateName") or "",
                "webpage_description": webpage_desc,
                # IMPORTANT: store stakeholder_id (not URI)
                "webpage_publisher": publisher_sid,
                "webpage_date_published": wp.get("datePublished") or "",
                "webpage_date_modified": wp.get("dateModified") or "",
                "webpage_main_entity_type": main_entity_type,
                # IMPORTANT: store stakeholder_id (not URI)
                "webpage_provider": provider_sid,
                # IMPORTANT: store website_id (not URI)
                "webpage_is_part_of": website_id,
            }

            # 4) Terms from about
            about_list = safe_list(wp.get("about"))
            for about in about_list:
                if not isinstance(about, dict):
                    continue
                term_name = about.get("name") or ""
                wd_url = about.get("sameAs") or ""
                if not term_name:
                    continue
                tid = term_id_for(reg, term_name, wd_url)
                terms_out[tid] = {
                    "term_id": tid,
                    "term_to_define": term_name,
                    "wikidata_url": wd_url,
                    "internal_definition": "",
                }
                wp_term_rows.append({
                    "webpage_id": webpage_id,
                    "webpage_is_relevant_to": tid,
                    "term_relevance_weight": 1.0,
                })

            # 5) Disclaimers: usageInfo.text + "disclaimer" headline content
            usage = wp.get("usageInfo") or {}
            if isinstance(usage, dict):
                usage_text = usage.get("text") or ""
                if usage_text:
                    did = disclaimer_id_for(reg, usage_text)
                    if did:
                        disclaimers_out[did] = {"disclaimer_id": did, "disclaimer_text": usage_text}
                        junct_disclaimer_rows.append({"webpage_id": webpage_id, "disclaimer_id": did})

            # Scoped nodes
            wp_id_uri = wp.get("@id")

            # 6) Main content sections
            has_part = safe_list(wp.get("hasPart"))
            has_part_ids = [hp.get("@id") for hp in has_part if isinstance(hp, dict) and hp.get("@id")]
            element_nodes = [
                n for n in nodes
                if n.get("@type") == "WebPageElement"
                and isinstance(n.get("isPartOf"), dict)
                and n["isPartOf"].get("@id") == wp_id_uri
            ]
            order_map = {hid: i + 1 for i, hid in enumerate(has_part_ids)}
            element_nodes.sort(key=lambda n: order_map.get(n.get("@id"), 10**9))

            display_order = 1
            for el in element_nodes:
                text = el.get("text") or ""
                el_id = el.get("@id") or ""
                if not text:
                    continue
                main_rows.append({
                    "webpage_id": webpage_id,
                    "main_content_text": text,
                    "main_content_text_id": el_id,
                    "display_order": order_map.get(el_id, display_order),
                })
                display_order += 1

                headline = (el.get("headline") or "").lower()
                if "disclaimer" in headline and text:
                    did = disclaimer_id_for(reg, text)
                    if did:
                        disclaimers_out[did] = {"disclaimer_id": did, "disclaimer_text": text}
                        junct_disclaimer_rows.append({"webpage_id": webpage_id, "disclaimer_id": did})

            # 7) FAQ rows
            faq_pages = [
                n for n in nodes
                if n.get("@type") == "FAQPage"
                and isinstance(n.get("isPartOf"), dict)
                and n["isPartOf"].get("@id") == wp_id_uri
            ]
            for faq_page in faq_pages:
                faq_id = faq_page.get("@id") or ""
                q_list = safe_list(faq_page.get("mainEntity"))
                q_order = 1
                for q in q_list:
                    if not isinstance(q, dict) or q.get("@type") != "Question":
                        continue
                    qid = q.get("@id") or ""
                    qtext = q.get("name") or ""
                    ans = q.get("acceptedAnswer") or {}
                    aid = ans.get("@id") if isinstance(ans, dict) else ""
                    atext = ans.get("text") if isinstance(ans, dict) else ""
                    if not qtext and not atext:
                        continue
                    faq_rows.append({
                        "webpage_id": webpage_id,
                        "faq_id": faq_id,
                        "faq_display_order": q_order,
                        "question_id": qid,
                        "question_text": qtext,
                        "answer_id": aid,
                        "answer_text": atext,
                    })
                    q_order += 1

            # 8) Webpage links: relatedLink + mentions.@id
            outbound_urls: List[str] = []
            outbound_urls.extend([u for u in safe_list(wp.get("relatedLink")) if isinstance(u, str)])
            for m in safe_list(wp.get("mentions")):
                if isinstance(m, dict) and m.get("@id"):
                    outbound_urls.append(m.get("@id"))

            # De-dupe + clean
            seen = set()
            for out in outbound_urls:
                if not out:
                    continue
                out_abs = to_internal_absolute(out)  # safe for external too (returns original if absolute)
                if not out_abs:
                    continue

                outn = norm_url(out_abs)
                if not outn or outn in seen:
                    continue
                seen.add(outn)

                # Classify internal vs external after normalisation
                if is_internal_ipfr_url(outn):
                    # Try resolve to a known webpage_id
                    dest_id = resolve_internal_destination_id(resolvers, outn)
                    if dest_id:
                        link_rows.append({
                            "webpage_id": webpage_id,
                            "destination_webpage_id": dest_id,
                            "destination_external_ref_id": "",
                        })
                    else:
                        # Internal but unresolved: do NOT create extref (keep external refs "pure")
                        unresolved_internal.append((webpage_id, outn))
                        link_rows.append({
                            "webpage_id": webpage_id,
                            "destination_webpage_id": "",
                            "destination_external_ref_id": "",
                        })
                else:
                    # External: create external reference
                    ext_id = external_ref_id_for(reg, outn)
                    host = urlparse(outn).netloc
                    title = host or outn
                    extrefs_out[ext_id] = {
                        "external_ref_id": ext_id,
                        "destination_link_title": title,
                        "destination_link_url": outn,
                    }
                    link_rows.append({
                        "webpage_id": webpage_id,
                        "destination_webpage_id": "",
                        "destination_external_ref_id": ext_id,
                    })

    # -----------------------------
    # Write outputs to sheets
    # -----------------------------

    # Stakeholders
    for sid in sorted(stakeholders_out.keys()):
        append_row(ws_stakeholder, H_STAKEHOLDER, stakeholders_out[sid])

    # Websites
    for wid in sorted(websites_out.keys()):
        append_row(ws_website, H_WEBSITE, websites_out[wid])

    # Webpages
    for wpid in sorted(webpages_out.keys()):
        append_row(ws_webpage, H_WEBPAGE, webpages_out[wpid])

    # External references
    for eid in sorted(extrefs_out.keys()):
        append_row(ws_extref, H_EXTREF, extrefs_out[eid])

    # Links (junction) – de-dupe and skip bad rows
    seen_links = set()
    for row in link_rows:
        src = (row.get("webpage_id") or "").strip()
        if not src:
            continue  # prevents blank webpage_id rows

        dwp = (row.get("destination_webpage_id") or "").strip()
        der = (row.get("destination_external_ref_id") or "").strip()

        # Enforce exactly one destination if resolved; allow both blank only for internal-unresolved
        if dwp and der:
            # Should never happen; prefer internal resolution if it does.
            der = ""
            row["destination_external_ref_id"] = ""

        key = (src, dwp, der)
        if key in seen_links:
            continue
        seen_links.add(key)

        append_row(ws_links, H_LINKS, row)

    # Main content
    for row in main_rows:
        append_row(ws_main, H_MAIN, row)

    # FAQ
    for row in faq_rows:
        append_row(ws_faq, H_FAQ, row)

    # Disclaimers
    for did in sorted(disclaimers_out.keys()):
        append_row(ws_disclaimer, H_DISCLAIMER, disclaimers_out[did])

    # Junction disclaimer (de-dupe)
    seen_pairs = set()
    for row in junct_disclaimer_rows:
        key = (row.get("webpage_id"), row.get("disclaimer_id"))
        if key in seen_pairs:
            continue
        seen_pairs.add(key)
        append_row(ws_jdisclaimer, H_JDISCLAIM, row)

    # Terms
    for tid in sorted(terms_out.keys()):
        append_row(ws_term, H_TERM, terms_out[tid])

    # Webpage->term junction (dedupe)
    seen_wpt = set()
    for row in wp_term_rows:
        key = (row.get("webpage_id"), row.get("webpage_is_relevant_to"))
        if key in seen_wpt:
            continue
        seen_wpt.add(key)
        append_row(ws_wprel, H_WPREL, row)

    # Term aliases (left empty unless you later add extraction)

    wb.save(xlsx_path)

    # Optional report
    internal_resolved = sum(1 for _, dwp, der in seen_links if dwp and not der)
    external_written = sum(1 for _, dwp, der in seen_links if der and not dwp)
    internal_unresolved = len(unresolved_internal)

    if report_path:
        try:
            # If report path is relative, anchor it next to the XLSX for convenience
            if not os.path.isabs(report_path):
                report_path = os.path.join(os.path.dirname(os.path.abspath(xlsx_path)), report_path)
            # Normalise extension to .jsonl
            if report_path.endswith(".csv"):
                report_path = report_path[:-4] + ".jsonl"

            import json as _json
            with open(report_path, "w", encoding="utf-8") as f:
                metrics = {
                    "webpages_total": len(webpages_out),
                    "links_internal_resolved_rows": internal_resolved,
                    "links_external_rows": external_written,
                    "links_internal_unresolved_rows": internal_unresolved,
                }
                f.write(_json.dumps(metrics, ensure_ascii=False) + "\n")
                for src, url in unresolved_internal:
                    f.write(_json.dumps({"source_webpage_id": src, "unresolved_internal_url": url}, ensure_ascii=False) + "\n")
        except Exception as e:
            print(f"⚠️ Could not write link report to {report_path}: {e}")

    # Basic stats output for logs
    print("✅ Completed JSON->XLSX load")
    print(f"   JSON dir:  {json_dir}")
    print(f"   XLSX:      {xlsx_path}")
    print(f"   WebPages:  {len(webpages_out)}")
    print(f"   Content:   {len(main_rows)}")
    print(f"   FAQs:      {len(faq_rows)}")
    print(f"   Ext refs:  {len(extrefs_out)}")
    print(f"   Terms:     {len(terms_out)}")
    print(f"   Orgs:      {len(stakeholders_out)}")
    print(f"   Links:     {len(seen_links)}")
    print(f"     - internal resolved:   {internal_resolved}")
    print(f"     - internal unresolved: {internal_unresolved}")
    print(f"     - external:            {external_written}")
    if report_path:
        print(f"   Link report: {report_path}")


if __name__ == "__main__":
    main()

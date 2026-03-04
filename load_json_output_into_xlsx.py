#!/usr/bin/env python3
"""
Load JSON-LD files from json_output/ into the existing XLSX "SQLite draft" workbook.

Design goals:
- Robust to variation in JSON-LD structure (WebPage + Article, WebPage + Service, GovernmentService, etc.)
- Deterministic IDs for entities that don't have a convenient identifier in the JSON (external refs, stakeholders, etc.)
- Only writes what is present in JSON files; vector embeddings / influence mapping etc. are left for manual filling later.

Sheets populated (if present in workbook):
- ENT_website
- ENT_stakeholder
- ENT_webpage
- ENT_webpage_main_content
- ENT_webpage_faq
- ENT_external_reference
- JUNCT_webpage_links
- ENT_term
- JUNCT_webpage_relevant_terms
- ENT_disclaimer
- JUNCT_disclaimer

Usage:
  python load_json_output_into_xlsx.py --json_dir json_output --xlsx_path ./260305_DB-Strcuture_05.xlsx

Notes:
- The script clears existing rows (row >= 2) in the target sheets it manages.
- It does NOT touch ENT_vector / influenced_by tables (by design).
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl


# -----------------------------
# Helpers
# -----------------------------

def sha1_hex(s: str) -> str:
    return hashlib.sha1(s.encode("utf-8")).hexdigest()

def stable_id(prefix: str, *parts: str, n: int = 10) -> str:
    joined = "||".join([p for p in parts if p is not None])
    return f"{prefix}{sha1_hex(joined)[:n]}"

def as_list(v: Any) -> List[Any]:
    if v is None:
        return []
    if isinstance(v, list):
        return v
    return [v]

def first_str(*vals: Any) -> str:
    for v in vals:
        if v is None:
            continue
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""

def strip_htmlish(text: str) -> str:
    # Keep it conservative; your content is mostly plaintext already.
    return re.sub(r"\s+", " ", (text or "")).strip()

def looks_like_internal_ipfr_url(url: str) -> bool:
    return "ipfirstresponse.ipaustralia.gov.au" in (url or "")

def slug_title_from_url(url: str) -> str:
    if not url:
        return ""
    try:
        path = re.sub(r"https?://", "", url).split("?", 1)[0].split("#", 1)[0]
        slug = path.rstrip("/").split("/")[-1]
        slug = slug.replace("-", " ").replace("_", " ").strip()
        return slug[:1].upper() + slug[1:] if slug else url
    except Exception:
        return url

def ensure_sheet(wb: openpyxl.Workbook, name: str) -> Optional[openpyxl.worksheet.worksheet.Worksheet]:
    return wb[name] if name in wb.sheetnames else None

def get_headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> List[str]:
    # First non-empty row is the header
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        if row and any(c is not None and str(c).strip() != "" for c in row):
            return [str(c).strip() if c is not None else "" for c in row]
    return []

def clear_data_rows(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    # Delete all rows after header row (assumes header is row 1)
    max_row = ws.max_row
    if max_row and max_row >= 2:
        ws.delete_rows(2, max_row - 1)

def append_row(ws: openpyxl.worksheet.worksheet.Worksheet, headers: List[str], row_dict: Dict[str, Any]) -> None:
    row = []
    for h in headers:
        if not h:
            row.append(None)
        else:
            row.append(row_dict.get(h, None))
    ws.append(row)


# -----------------------------
# JSON-LD extraction
# -----------------------------

@dataclass
class Extracted:
    # ENT_website
    websites: List[Dict[str, Any]]
    # ENT_stakeholder
    stakeholders: List[Dict[str, Any]]
    # ENT_webpage
    webpages: List[Dict[str, Any]]
    # ENT_webpage_main_content
    main_content: List[Dict[str, Any]]
    # ENT_webpage_faq
    faqs: List[Dict[str, Any]]
    # ENT_external_reference
    external_refs: List[Dict[str, Any]]
    # JUNCT_webpage_links
    webpage_links: List[Dict[str, Any]]
    # ENT_term
    terms: List[Dict[str, Any]]
    # JUNCT_webpage_relevant_terms
    webpage_terms: List[Dict[str, Any]]
    # ENT_disclaimer
    disclaimers: List[Dict[str, Any]]
    # JUNCT_disclaimer
    webpage_disclaimers: List[Dict[str, Any]]


def index_graph(graph: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    idx: Dict[str, Dict[str, Any]] = {}
    for node in graph:
        nid = node.get("@id")
        if isinstance(nid, str) and nid.strip():
            idx[nid.strip()] = node
    return idx


def node_types(node: Dict[str, Any]) -> List[str]:
    t = node.get("@type")
    if isinstance(t, list):
        return [str(x) for x in t]
    if isinstance(t, str):
        return [t]
    return []


def find_first(graph: List[Dict[str, Any]], typ: str) -> Optional[Dict[str, Any]]:
    for n in graph:
        if typ in node_types(n):
            return n
    return None


def find_all(graph: List[Dict[str, Any]], typ: str) -> List[Dict[str, Any]]:
    return [n for n in graph if typ in node_types(n)]


def extract_from_file(path: Path) -> Extracted:
    raw = json.loads(path.read_text(encoding="utf-8"))
    graph = raw.get("@graph", [])
    if not isinstance(graph, list):
        graph = []

    idx = index_graph(graph)

    websites: List[Dict[str, Any]] = []
    stakeholders: List[Dict[str, Any]] = []
    webpages: List[Dict[str, Any]] = []
    main_content: List[Dict[str, Any]] = []
    faqs: List[Dict[str, Any]] = []
    external_refs: List[Dict[str, Any]] = []
    webpage_links: List[Dict[str, Any]] = []
    terms: List[Dict[str, Any]] = []
    webpage_terms: List[Dict[str, Any]] = []
    disclaimers: List[Dict[str, Any]] = []
    webpage_disclaimers: List[Dict[str, Any]] = []

    # WebSite
    ws_node = find_first(graph, "WebSite")
    if ws_node:
        website_uri = first_str(ws_node.get("@id"))
        website_url = first_str(ws_node.get("url"))
        website_title = first_str(ws_node.get("name"))
        publisher = ws_node.get("publisher", {})
        owned_by = ""
        if isinstance(publisher, dict):
            owned_by = first_str(publisher.get("@id"), publisher.get("name"))
        website_id = stable_id("site_", website_uri or website_url or website_title)

        websites.append(
            {
                "website_id": website_id,
                "website_uri": website_uri,
                "website_title": website_title,
                "website_url": website_url,
                "website_owned_by": owned_by,
            }
        )

    # Stakeholders: Organization-ish nodes
    org_types = {"Organization", "GovernmentOrganization"}
    for node in graph:
        if not org_types.intersection(set(node_types(node))):
            continue

        uri = first_str(node.get("@id"))
        name = first_str(node.get("name"))
        url = first_str(node.get("url"))
        description = first_str(node.get("description"))
        same_as = node.get("sameAs")

        # sameAs may be list or str
        same_as_str = ""
        if isinstance(same_as, list):
            same_as_str = "; ".join([str(x) for x in same_as if x])
        elif isinstance(same_as, str):
            same_as_str = same_as

        stakeholder_id = stable_id("stake_", uri or name or url)

        stakeholders.append(
            {
                "stakeholder_id": stakeholder_id,
                "stakeholder_uri": uri,
                "stakeholder_type": ", ".join(node_types(node)),
                "stakeholder_name": name,
                "stakeholder_alias": first_str(node.get("alternateName")),
                "stakeholder_description": description,
                "stakeholder_url": url,
                "stakeholder_same_as": same_as_str,
            }
        )

    # WebPage (this is the main entry per file)
    wp_node = find_first(graph, "WebPage")
    if not wp_node:
        # If there's no WebPage, still return empty; nothing else to do safely.
        return Extracted(
            websites=websites,
            stakeholders=stakeholders,
            webpages=webpages,
            main_content=main_content,
            faqs=faqs,
            external_refs=external_refs,
            webpage_links=webpage_links,
            terms=terms,
            webpage_terms=webpage_terms,
            disclaimers=disclaimers,
            webpage_disclaimers=webpage_disclaimers,
        )

    webpage_id = first_str(wp_node.get("identifier"))
    webpage_uri = first_str(wp_node.get("@id"))
    webpage_url = first_str(wp_node.get("url"))
    webpage_name = first_str(wp_node.get("name"))
    webpage_altname = first_str(wp_node.get("alternateName"))
    webpage_description = first_str(wp_node.get("description"))
    webpage_date_published = first_str(wp_node.get("datePublished"))
    webpage_date_modified = first_str(wp_node.get("dateModified"))

    publisher = wp_node.get("publisher", {})
    webpage_publisher = ""
    if isinstance(publisher, dict):
        webpage_publisher = first_str(publisher.get("@id"), publisher.get("name"))

    is_part_of = wp_node.get("isPartOf", {})
    webpage_is_part_of = ""
    if isinstance(is_part_of, dict):
        webpage_is_part_of = first_str(is_part_of.get("@id"), is_part_of.get("name"))

    # Identify the main entity and its type (Article vs Service vs GovernmentService)
    main_entity = wp_node.get("mainEntity")
    main_entity_id = ""
    if isinstance(main_entity, dict):
        main_entity_id = first_str(main_entity.get("@id"))
    elif isinstance(main_entity, str):
        main_entity_id = main_entity

    main_entity_node = idx.get(main_entity_id, {}) if main_entity_id else {}
    main_entity_types = node_types(main_entity_node) if main_entity_node else []
    webpage_main_entity_type = ", ".join(main_entity_types)

    # Provider: for Service/GovernmentService, provider/serviceOperator often points to org @id
    webpage_provider = ""
    if main_entity_node:
        provider = main_entity_node.get("provider") or main_entity_node.get("serviceOperator")
        if isinstance(provider, dict):
            webpage_provider = first_str(provider.get("@id"), provider.get("name"))
        elif isinstance(provider, str):
            webpage_provider = provider

    webpages.append(
        {
            "webpage_id": webpage_id,
            "webpage_uri": webpage_uri,
            "webpage_url": webpage_url,
            "webpage_name": webpage_name,
            "webpage_altname": webpage_altname,
            "webpage_description": webpage_description,
            "webpage_publisher": webpage_publisher,
            "webpage_date_published": webpage_date_published,
            "webpage_date_modified": webpage_date_modified,
            "webpage_main_entity_type": webpage_main_entity_type,
            "webpage_provider": webpage_provider,
            "webpage_is_part_of": webpage_is_part_of,
        }
    )

    # Disclaimer (usageInfo.text on WebPage)
    usage_info = wp_node.get("usageInfo")
    if isinstance(usage_info, dict):
        disc_text = first_str(usage_info.get("text"))
        if disc_text:
            disclaimer_id = stable_id("disc_", disc_text)
            disclaimers.append(
                {"disclaimer_id": disclaimer_id, "disclaimer_text": disc_text}
            )
            webpage_disclaimers.append(
                {"webpage_id": webpage_id, "disclaimer_id": disclaimer_id}
            )

    # Terms ("about": Thing{name, sameAs})
    about_list = as_list(wp_node.get("about"))
    for thing in about_list:
        if not isinstance(thing, dict):
            continue
        tname = first_str(thing.get("name"))
        same_as = first_str(thing.get("sameAs"))
        if not tname:
            continue
        term_id = stable_id("term_", tname, same_as)
        terms.append(
            {
                "term_id": term_id,
                "term_to_define": tname,
                "wikidata_url": same_as,
                "internal_definition": "",
            }
        )
        webpage_terms.append(
            {
                "webpage_id": webpage_id,
                "webpage_is_relevant_to": term_id,
                "term_relevance_weight": 1.0,
            }
        )

    # Main content:
    # - articleBody for Article (if present)
    # - WebPageElement nodes (ordered by "position" where available)
    if "Article" in main_entity_types:
        article_body = first_str(main_entity_node.get("articleBody"))
        if article_body:
            main_content.append(
                {
                    "webpage_id": webpage_id,
                    "main_content_text": article_body,
                    "main_content_text_id": stable_id("mc_", webpage_id, "articleBody"),
                    "display_order": 0,
                }
            )

    elements = find_all(graph, "WebPageElement")
    # Sort by numeric "position" if present; otherwise stable by @id.
    def elem_sort_key(n: Dict[str, Any]) -> Tuple[int, str]:
        pos = n.get("position")
        try:
            pos_i = int(pos)
        except Exception:
            pos_i = 10_000
        return (pos_i, first_str(n.get("@id")))

    elements_sorted = sorted(elements, key=elem_sort_key)

    # Exclude elements that look like the FAQ container; keep "Disclaimer" sections as content too
    # (disclaimer table is already handled via usageInfo, but the visible disclaimer section can still be content).
    order_base = 1  # after articleBody
    current_order = order_base

    for el in elements_sorted:
        el_id = first_str(el.get("@id"))
        headline = first_str(el.get("headline"))
        text = first_str(el.get("text"))
        if not text and not headline:
            continue

        # Skip if it is the FAQ itself (FAQPage handled separately)
        if "#faq" in el_id.lower():
            continue

        # Write a readable text chunk: headline + text (keeps context)
        chunk = ""
        if headline:
            chunk += headline.strip() + "\n\n"
        chunk += text.strip()

        main_content.append(
            {
                "webpage_id": webpage_id,
                "main_content_text": chunk,
                "main_content_text_id": el_id or stable_id("mc_", webpage_id, headline, text),
                "display_order": el.get("position") if el.get("position") is not None else current_order,
            }
        )
        current_order += 1

    # FAQ
    faq_page = find_first(graph, "FAQPage")
    if faq_page:
        faq_id = first_str(faq_page.get("@id")) or stable_id("faq_", webpage_id)
        q_list = as_list(faq_page.get("mainEntity"))
        display_order = 1
        for q in q_list:
            if not isinstance(q, dict):
                continue
            if first_str(q.get("@type")) != "Question":
                # Sometimes schema uses {"@type":"Question"}; if it's missing, still proceed.
                pass
            question_id = first_str(q.get("@id")) or stable_id("q_", faq_id, str(display_order))
            question_text = first_str(q.get("name"))

            ans = q.get("acceptedAnswer") or {}
            answer_id = ""
            answer_text = ""
            if isinstance(ans, dict):
                answer_id = first_str(ans.get("@id")) or stable_id("a_", question_id)
                answer_text = first_str(ans.get("text"))
            else:
                answer_id = stable_id("a_", question_id)
                answer_text = first_str(ans)

            if not (question_text or answer_text):
                continue

            faqs.append(
                {
                    "webpage_id": webpage_id,
                    "faq_id": faq_id,
                    "faq_display_order": display_order,
                    "question_id": question_id,
                    "question_text": question_text,
                    "answer_id": answer_id,
                    "answer_text": answer_text,
                }
            )
            display_order += 1

    # Links:
    # - relatedLink (list of URLs)
    # - mentions (often list of {"@id": "..."} or nodes)
    related_links = as_list(wp_node.get("relatedLink"))
    mentions = as_list(wp_node.get("mentions"))

    all_link_urls: List[str] = []
    for v in related_links:
        if isinstance(v, str):
            all_link_urls.append(v)
    for m in mentions:
        if isinstance(m, dict):
            mid = first_str(m.get("@id"))
            if mid.startswith("http"):
                all_link_urls.append(mid)

    # Dedup URLs while preserving order
    seen = set()
    dedup_urls: List[str] = []
    for u in all_link_urls:
        u2 = u.strip()
        if not u2 or u2 in seen:
            continue
        seen.add(u2)
        dedup_urls.append(u2)

    for url in dedup_urls:
        # If internal but we cannot resolve identifier here, treat as external reference for now.
        ext_id = stable_id("extref_", url)
        external_refs.append(
            {
                "external_ref_id": ext_id,
                "destination_link_title": slug_title_from_url(url),
                "destination_link_url": url,
            }
        )
        webpage_links.append(
            {
                "webpage_id": webpage_id,
                "destination_webpage_id": "",
                "destination_external_ref_id": ext_id,
            }
        )

    return Extracted(
        websites=websites,
        stakeholders=stakeholders,
        webpages=webpages,
        main_content=main_content,
        faqs=faqs,
        external_refs=external_refs,
        webpage_links=webpage_links,
        terms=terms,
        webpage_terms=webpage_terms,
        disclaimers=disclaimers,
        webpage_disclaimers=webpage_disclaimers,
    )


# -----------------------------
# Workbook writing
# -----------------------------

def dedupe_rows(rows: List[Dict[str, Any]], key_fields: List[str]) -> List[Dict[str, Any]]:
    seen = set()
    out = []
    for r in rows:
        k = tuple((r.get(f) or "") for f in key_fields)
        if k in seen:
            continue
        seen.add(k)
        out.append(r)
    return out


def load_all(json_dir: Path) -> Extracted:
    websites: List[Dict[str, Any]] = []
    stakeholders: List[Dict[str, Any]] = []
    webpages: List[Dict[str, Any]] = []
    main_content: List[Dict[str, Any]] = []
    faqs: List[Dict[str, Any]] = []
    external_refs: List[Dict[str, Any]] = []
    webpage_links: List[Dict[str, Any]] = []
    terms: List[Dict[str, Any]] = []
    webpage_terms: List[Dict[str, Any]] = []
    disclaimers: List[Dict[str, Any]] = []
    webpage_disclaimers: List[Dict[str, Any]] = []

    for p in sorted(json_dir.glob("*.json")):
        ex = extract_from_file(p)
        websites.extend(ex.websites)
        stakeholders.extend(ex.stakeholders)
        webpages.extend(ex.webpages)
        main_content.extend(ex.main_content)
        faqs.extend(ex.faqs)
        external_refs.extend(ex.external_refs)
        webpage_links.extend(ex.webpage_links)
        terms.extend(ex.terms)
        webpage_terms.extend(ex.webpage_terms)
        disclaimers.extend(ex.disclaimers)
        webpage_disclaimers.extend(ex.webpage_disclaimers)

    # Dedupe by stable keys
    websites = dedupe_rows(websites, ["website_uri", "website_url"])
    stakeholders = dedupe_rows(stakeholders, ["stakeholder_uri", "stakeholder_name"])
    webpages = dedupe_rows(webpages, ["webpage_id"])
    external_refs = dedupe_rows(external_refs, ["destination_link_url"])
    terms = dedupe_rows(terms, ["term_to_define", "wikidata_url"])
    disclaimers = dedupe_rows(disclaimers, ["disclaimer_id"])

    # Junction tables: dedupe by (webpage_id, target)
    webpage_links = dedupe_rows(webpage_links, ["webpage_id", "destination_webpage_id", "destination_external_ref_id"])
    webpage_terms = dedupe_rows(webpage_terms, ["webpage_id", "webpage_is_relevant_to"])
    webpage_disclaimers = dedupe_rows(webpage_disclaimers, ["webpage_id", "disclaimer_id"])

    return Extracted(
        websites=websites,
        stakeholders=stakeholders,
        webpages=webpages,
        main_content=main_content,
        faqs=faqs,
        external_refs=external_refs,
        webpage_links=webpage_links,
        terms=terms,
        webpage_terms=webpage_terms,
        disclaimers=disclaimers,
        webpage_disclaimers=webpage_disclaimers,
    )


def write_to_workbook(xlsx_path: Path, extracted: Extracted) -> None:
    wb = openpyxl.load_workbook(xlsx_path)

    # Map sheet -> rows
    sheet_rows: Dict[str, List[Dict[str, Any]]] = {
        "ENT_website": extracted.websites,
        "ENT_stakeholder": extracted.stakeholders,
        "ENT_webpage": extracted.webpages,
        "ENT_webpage_main_content": extracted.main_content,
        "ENT_webpage_faq": extracted.faqs,
        "ENT_external_reference": extracted.external_refs,
        "JUNCT_webpage_links": extracted.webpage_links,
        "ENT_term": extracted.terms,
        "JUNCT_webpage_relevant_terms": extracted.webpage_terms,
        "ENT_disclaimer": extracted.disclaimers,
        "JUNCT_disclaimer": extracted.webpage_disclaimers,
    }

    for sheet_name, rows in sheet_rows.items():
        ws = ensure_sheet(wb, sheet_name)
        if ws is None:
            # Workbook may not include every optional sheet; skip quietly.
            continue

        headers = get_headers(ws)
        if not headers:
            continue

        clear_data_rows(ws)

        # Some sheets have trailing empty header columns; append_row handles those.
        for r in rows:
            append_row(ws, headers, r)

    wb.save(xlsx_path)


# -----------------------------
# CLI
# -----------------------------

def guess_repo_root(start: Path) -> Path:
    """
    Walk upwards looking for a directory that contains:
      - json_output/  (dir)
      - at least one *.xlsx
    """
    cur = start.resolve()
    for _ in range(8):
        if (cur / "json_output").is_dir() and any(cur.glob("*.xlsx")):
            return cur
        cur = cur.parent
    return start.resolve()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--json_dir", default="json_output", help="Folder containing JSON files (default: json_output)")
    parser.add_argument("--xlsx_path", default="", help="Path to the XLSX file to populate (default: auto-detect in repo root)")
    args = parser.parse_args()

    here = Path(__file__).resolve()
    repo_root = guess_repo_root(here.parent)

    json_dir = (repo_root / args.json_dir).resolve() if not Path(args.json_dir).is_absolute() else Path(args.json_dir).resolve()
    if not json_dir.is_dir():
        raise SystemExit(f"ERROR: json_dir not found: {json_dir}")

    if args.xlsx_path:
        xlsx_path = (repo_root / args.xlsx_path).resolve() if not Path(args.xlsx_path).is_absolute() else Path(args.xlsx_path).resolve()
    else:
        # pick the first XLSX in repo root (stable sort)
        xlsx_candidates = sorted(repo_root.glob("*.xlsx"))
        if not xlsx_candidates:
            raise SystemExit(f"ERROR: no .xlsx found in repo root: {repo_root}")
        xlsx_path = xlsx_candidates[0]

    if not xlsx_path.exists():
        raise SystemExit(f"ERROR: xlsx_path not found: {xlsx_path}")

    extracted = load_all(json_dir)
    write_to_workbook(xlsx_path, extracted)

    print("✅ Completed JSON->XLSX load")
    print(f"   Repo root: {repo_root}")
    print(f"   JSON dir:  {json_dir}")
    print(f"   XLSX:      {xlsx_path}")
    print(f"   WebPages:  {len(extracted.webpages)}")
    print(f"   Content:   {len(extracted.main_content)}")
    print(f"   FAQs:      {len(extracted.faqs)}")
    print(f"   Ext refs:  {len(extracted.external_refs)}")
    print(f"   Terms:     {len(extracted.terms)}")
    print(f"   Orgs:      {len(extracted.stakeholders)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
